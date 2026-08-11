"""Validate and normalize the AthleteOS workout workbook into a runtime JSON asset."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl


SHEET_KEYS = {
    "Split_Templates": "split_id",
    "Program_Days": "day_template_id",
    "Day_Exercises": "prescription_id",
    "Exercise_Catalog": "exercise_id",
    "Progression_Rules": "rule_id",
    "Selection_Rules": "rule_id",
    "Substitutions": "group_id",
    "Research_Sources": "source_id",
}


def sheet_rows(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    if not all(isinstance(header, str) and header for header in headers):
        raise ValueError(f"{sheet_name} has an invalid header row")
    rows = []
    key = SHEET_KEYS[sheet_name]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        padded = (*values, *([None] * max(0, len(headers) - len(values))))
        row = dict(zip(headers, padded[: len(headers)], strict=True))
        if row.get(key) is None:
            continue
        if sheet_name == "Research_Sources" and not str(row[key]).startswith("SRC"):
            continue
        rows.append(row)
    ids = [row[key] for row in rows]
    if len(ids) != len(set(ids)):
        raise ValueError(f"{sheet_name} contains duplicate {key} values")
    return rows


def validate(data: dict[str, list[dict[str, Any]]]) -> None:
    split_ids = {row["split_id"] for row in data["Split_Templates"]}
    day_ids = {row["day_template_id"] for row in data["Program_Days"]}
    exercise_ids = {row["exercise_id"] for row in data["Exercise_Catalog"]}
    progression_ids = {row["rule_id"] for row in data["Progression_Rules"]}
    substitution_ids = {row["group_id"] for row in data["Substitutions"]}

    for row in data["Program_Days"]:
        if row["split_id"] not in split_ids:
            raise ValueError(f"Unknown split_id in Program_Days: {row['split_id']}")
    for row in data["Day_Exercises"]:
        if row["day_template_id"] not in day_ids:
            raise ValueError(f"Unknown day_template_id: {row['day_template_id']}")
        if row["exercise_id"] not in exercise_ids:
            raise ValueError(f"Unknown exercise_id: {row['exercise_id']}")
        if row["progression_rule_id"] not in progression_ids:
            raise ValueError(f"Unknown progression_rule_id: {row['progression_rule_id']}")
        if row["substitution_group_id"] and row["substitution_group_id"] not in substitution_ids:
            raise ValueError(f"Unknown substitution_group_id: {row['substitution_group_id']}")

    exercise_names = {row["name"].casefold() for row in data["Exercise_Catalog"]}
    for row in data["Substitutions"]:
        names = [row["default_exercise"], *str(row["alternatives"]).split(";")]
        unknown = [name.strip() for name in names if name.strip().casefold() not in exercise_names]
        if unknown:
            raise ValueError(f"Unknown substitution exercise names in {row['group_id']}: {unknown}")


def build(source: Path, destination: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    missing = set(SHEET_KEYS) - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing sheets: {sorted(missing)}")
    sheets = {sheet: sheet_rows(workbook, sheet) for sheet in SHEET_KEYS}
    validate(sheets)
    payload = {
        "metadata": {
            "name": "AthleteOS Gym Workout Dataset",
            "version": "1.0",
            "generated": "2026-08-11",
            "source_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
            "counts": {sheet: len(rows) for sheet, rows in sheets.items()},
        },
        "sheets": sheets,
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    args = parser.parse_args()
    payload = build(args.source.resolve(), args.destination.resolve())
    print(json.dumps(payload["metadata"], indent=2))


if __name__ == "__main__":
    main()
